from openpyxl import load_workbook
import json

ARQUIVO_EXCEL = "GENERO_11012026.xlsx"
ARQUIVO_JSON = "musicas.json"

print("Abrindo planilha...")

wb = load_workbook(
    ARQUIVO_EXCEL,
    read_only=True,
    data_only=True
)

ws = wb.active

headers = [
    str(c.value).strip()
    if c.value is not None
    else ""
    for c in next(ws.rows)
]

idx_cod = headers.index("Cod")
idx_titulo = headers.index("Título")
idx_interprete = headers.index("Intérprete")
idx_genero = headers.index("Gênero")
idx_idioma = headers.index("Idioma")

musicas = []

for row in ws.iter_rows(
    min_row=2,
    values_only=True
):

    codigo = row[idx_cod]
    titulo = row[idx_titulo]
    interprete = row[idx_interprete]
    genero = row[idx_genero]
    idioma = row[idx_idioma]

    if codigo is None:
        continue

    if titulo is None:
        continue

    musicas.append({

        "codigo":
        str(codigo).strip(),

        "musica":
        str(titulo).strip(),

        "artista":
        ""
        if interprete is None
        else str(interprete).strip(),

        "genero":
        ""
        if genero is None
        else str(genero).strip(),

        "idioma":
        ""
        if idioma is None
        else str(idioma).strip()

    })

musicas.sort(
    key=lambda x:
    x["musica"].upper()
)

with open(
    ARQUIVO_JSON,
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        musicas,
        f,
        ensure_ascii=False,
        indent=2
    )

print()
print("JSON gerado com sucesso!")
print(f"Total de músicas: {len(musicas)}")
print(f"Arquivo: {ARQUIVO_JSON}")